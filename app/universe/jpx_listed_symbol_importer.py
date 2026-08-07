"""JPX公式の東証上場銘柄一覧をlisted_symbolsへ取り込む。"""

from __future__ import annotations

import csv
import html
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from html.parser import HTMLParser
from pathlib import Path
from typing import Callable, Iterable
from urllib.parse import urljoin
from urllib.request import Request, urlopen

from app.universe.listed_symbol_repository import (
    ListedSymbolRepository,
)
from app.universe.universe_models import ListedSymbol


DEFAULT_JPX_PAGE_URL = (
    "https://www.jpx.co.jp/markets/"
    "statistics-equities/misc/01.html"
)
DEFAULT_SOURCE_NAME = "jpx-listed-issues"
_ALLOWED_MARKETS = {"Prime", "Standard", "Growth"}


@dataclass(frozen=True, slots=True)
class JpxListedSymbolImportResult:
    """JPX銘柄マスター取込結果。"""

    generated_at: datetime
    source_url: str | None
    source_path: str
    raw_row_count: int
    imported_count: int
    skipped_count: int
    prime_count: int
    standard_count: int
    growth_count: int

    def to_dict(self) -> dict[str, object]:
        return {
            "generated_at": self.generated_at.isoformat(),
            "source_url": self.source_url,
            "source_path": self.source_path,
            "raw_row_count": self.raw_row_count,
            "imported_count": self.imported_count,
            "skipped_count": self.skipped_count,
            "prime_count": self.prime_count,
            "standard_count": self.standard_count,
            "growth_count": self.growth_count,
        }


class _AnchorCollector(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self._href: str | None = None
        self._text_parts: list[str] = []
        self.anchors: list[tuple[str, str]] = []

    def handle_starttag(self, tag: str, attrs) -> None:
        if tag.lower() != "a":
            return
        values = dict(attrs)
        self._href = values.get("href")
        self._text_parts = []

    def handle_data(self, data: str) -> None:
        if self._href is not None:
            self._text_parts.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() != "a" or self._href is None:
            return

        text = " ".join(
            part.strip()
            for part in self._text_parts
            if part.strip()
        )
        self.anchors.append(
            (
                self._href,
                html.unescape(text),
            )
        )
        self._href = None
        self._text_parts = []


class JpxListedSymbolImporter:
    """JPX月次上場銘柄一覧を取得・正規化・保存する。"""

    def __init__(
        self,
        *,
        database_path: Path,
        source_name: str = DEFAULT_SOURCE_NAME,
        now_provider: Callable[[], datetime] | None = None,
        timeout_seconds: float = 30.0,
        user_agent: str = "Project-KATANA/1.0",
    ) -> None:
        if timeout_seconds <= 0:
            raise ValueError(
                "timeout_secondsは0より大きい必要があります。"
            )

        self.database_path = Path(database_path)
        self.repository = ListedSymbolRepository(
            self.database_path
        )
        self.source_name = source_name.strip()
        self.now_provider = (
            now_provider
            if now_provider is not None
            else lambda: datetime.now(timezone.utc)
        )
        self.timeout_seconds = float(timeout_seconds)
        self.user_agent = user_agent

        if not self.source_name:
            raise ValueError(
                "source_nameを指定してください。"
            )

    def download_latest(
        self,
        *,
        page_url: str = DEFAULT_JPX_PAGE_URL,
        destination_directory: Path = Path(
            "data/reference/jpx"
        ),
    ) -> tuple[Path, str]:
        """JPX公式ページから最新Excelリンクを検出して保存する。"""

        page = self._download_bytes(page_url).decode(
            "utf-8",
            errors="replace",
        )
        workbook_url = self._discover_workbook_url(
            page_url=page_url,
            page_html=page,
        )

        suffix = (
            ".xlsx"
            if workbook_url.lower().endswith(".xlsx")
            else ".xls"
        )
        destination_directory = Path(
            destination_directory
        )
        destination_directory.mkdir(
            parents=True,
            exist_ok=True,
        )
        path = (
            destination_directory
            / f"tse_listed_issues_latest{suffix}"
        )
        temporary = path.with_suffix(
            path.suffix + ".tmp"
        )
        temporary.write_bytes(
            self._download_bytes(workbook_url)
        )
        temporary.replace(path)
        return path, workbook_url

    def import_latest(
        self,
        *,
        page_url: str = DEFAULT_JPX_PAGE_URL,
        destination_directory: Path = Path(
            "data/reference/jpx"
        ),
    ) -> JpxListedSymbolImportResult:
        path, source_url = self.download_latest(
            page_url=page_url,
            destination_directory=destination_directory,
        )
        return self.import_file(
            path,
            source_url=source_url,
        )

    def import_file(
        self,
        path: Path,
        *,
        source_url: str | None = None,
    ) -> JpxListedSymbolImportResult:
        resolved = Path(path)
        if not resolved.exists():
            raise FileNotFoundError(resolved)

        rows = tuple(self._read_rows(resolved))
        now = self._current_time()

        symbols: list[ListedSymbol] = []
        skipped = 0

        for row in rows:
            symbol = self._to_symbol(
                row,
                updated_at=now,
            )
            if symbol is None:
                skipped += 1
                continue
            symbols.append(symbol)

        unique = {
            item.code: item
            for item in symbols
        }
        normalized = tuple(
            unique[code]
            for code in sorted(unique)
        )

        self.repository.replace_active_snapshot(
            normalized,
            source_name=self.source_name,
            snapshot_at=now,
        )

        market_counts = {
            market: sum(
                item.market == market
                for item in normalized
            )
            for market in _ALLOWED_MARKETS
        }

        return JpxListedSymbolImportResult(
            generated_at=now,
            source_url=source_url,
            source_path=str(resolved),
            raw_row_count=len(rows),
            imported_count=len(normalized),
            skipped_count=skipped,
            prime_count=market_counts["Prime"],
            standard_count=market_counts["Standard"],
            growth_count=market_counts["Growth"],
        )

    def _download_bytes(self, url: str) -> bytes:
        request = Request(
            url,
            headers={
                "User-Agent": self.user_agent,
                "Accept": (
                    "text/html,application/vnd.ms-excel,"
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet,*/*"
                ),
            },
        )
        with urlopen(
            request,
            timeout=self.timeout_seconds,
        ) as response:
            return response.read()

    @staticmethod
    def _discover_workbook_url(
        *,
        page_url: str,
        page_html: str,
    ) -> str:
        parser = _AnchorCollector()
        parser.feed(page_html)

        candidates: list[tuple[int, str]] = []

        for href, text in parser.anchors:
            lower_href = href.lower()
            if not (
                lower_href.endswith(".xls")
                or lower_href.endswith(".xlsx")
            ):
                continue

            normalized = f"{href} {text}".lower()
            score = 0

            if "data_j" in normalized:
                score += 100
            if "上場銘柄" in normalized:
                score += 80
            if "listed" in normalized:
                score += 50
            if "issues" in normalized:
                score += 30
            if "topix" in normalized:
                score -= 100

            candidates.append(
                (
                    score,
                    urljoin(page_url, href),
                )
            )

        if not candidates:
            raw_links = re.findall(
                r"href=[\"']([^\"']+\.(?:xlsx?|XLSX?))[\"']",
                page_html,
            )
            candidates = [
                (
                    100
                    if "data_j" in href.lower()
                    else 0,
                    urljoin(page_url, href),
                )
                for href in raw_links
                if "topix" not in href.lower()
            ]

        if not candidates:
            raise RuntimeError(
                "JPXページから上場銘柄一覧Excelを"
                "検出できませんでした。"
            )

        candidates.sort(
            key=lambda item: item[0],
            reverse=True,
        )
        return candidates[0][1]

    def _read_rows(
        self,
        path: Path,
    ) -> Iterable[dict[str, object]]:
        suffix = path.suffix.lower()

        if suffix == ".csv":
            yield from self._read_csv(path)
            return

        if suffix == ".xlsx":
            yield from self._read_xlsx(path)
            return

        if suffix == ".xls":
            yield from self._read_xls(path)
            return

        raise ValueError(
            "対応形式は.csv/.xls/.xlsxです。 "
            f"path={path}"
        )

    @staticmethod
    def _read_csv(
        path: Path,
    ) -> Iterable[dict[str, object]]:
        last_error: Exception | None = None

        for encoding in (
            "utf-8-sig",
            "cp932",
            "utf-8",
        ):
            try:
                with path.open(
                    "r",
                    encoding=encoding,
                    newline="",
                ) as handle:
                    rows = list(csv.DictReader(handle))
                yield from rows
                return
            except UnicodeError as error:
                last_error = error

        raise UnicodeError(
            "CSVの文字コードを判定できません。"
        ) from last_error

    @staticmethod
    def _read_xlsx(
        path: Path,
    ) -> Iterable[dict[str, object]]:
        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise RuntimeError(
                ".xlsx読込にはopenpyxlが必要です。 "
                "python -m pip install openpyxl"
            ) from error

        workbook = load_workbook(
            path,
            read_only=True,
            data_only=True,
        )
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            if header is None:
                return

            keys = [
                str(value).strip()
                if value is not None
                else ""
                for value in header
            ]

            for values in rows:
                yield {
                    keys[index]: values[index]
                    for index in range(
                        min(len(keys), len(values))
                    )
                    if keys[index]
                }
        finally:
            workbook.close()

    @staticmethod
    def _read_xls(
        path: Path,
    ) -> Iterable[dict[str, object]]:
        try:
            import xlrd
        except ImportError as error:
            raise RuntimeError(
                "JPXの.xls読込にはxlrdが必要です。 "
                "python -m pip install xlrd"
            ) from error

        workbook = xlrd.open_workbook(str(path))
        sheet = workbook.sheet_by_index(0)

        if sheet.nrows <= 0:
            return

        keys = [
            str(sheet.cell_value(0, column)).strip()
            for column in range(sheet.ncols)
        ]

        for row_index in range(1, sheet.nrows):
            yield {
                keys[column]: sheet.cell_value(
                    row_index,
                    column,
                )
                for column in range(sheet.ncols)
                if keys[column]
            }

    def _to_symbol(
        self,
        row: dict[str, object],
        *,
        updated_at: datetime,
    ) -> ListedSymbol | None:
        code = self._first(
            row,
            ("コード", "Code", "code"),
        )
        name = self._first(
            row,
            ("銘柄名", "Issue Name", "Name", "name"),
        )
        category = self._first(
            row,
            (
                "市場・商品区分",
                "市場区分",
                "Market/Products",
                "Market Segment",
                "market",
            ),
        )

        normalized_code = self._normalize_code(code)
        normalized_name = str(name or "").strip()
        market = self._normalize_market(category)

        if (
            normalized_code is None
            or not normalized_name
            or market is None
        ):
            return None

        category_text = str(category or "")
        if not self._is_domestic_common_stock(
            category_text
        ):
            return None

        return ListedSymbol(
            code=normalized_code,
            name=normalized_name,
            market=market,
            security_type="common_stock",
            trading_unit=100,
            listed_date=None,
            delisted_date=None,
            is_active=True,
            source=self.source_name,
            updated_at=updated_at,
        )

    @staticmethod
    def _first(
        row: dict[str, object],
        aliases: tuple[str, ...],
    ) -> object | None:
        normalized = {
            str(key).strip().lower(): value
            for key, value in row.items()
        }

        for alias in aliases:
            key = alias.strip().lower()
            if key in normalized:
                value = normalized[key]
                if value not in (None, ""):
                    return value
        return None

    @staticmethod
    def _normalize_code(
        value: object | None,
    ) -> str | None:
        if value is None:
            return None

        if isinstance(value, float):
            if not value.is_integer():
                return None
            text = str(int(value))
        else:
            text = str(value).strip()

        text = re.sub(r"\.0$", "", text)

        text = text.upper()

        # 東証の新しい証券コードには英字を含む4桁コード
        # （例: 607A）が存在するため、数字のみには限定しない。
        if not re.fullmatch(r"[0-9A-Z]{4}", text):
            return None

        return text

    @staticmethod
    def _normalize_market(
        value: object | None,
    ) -> str | None:
        text = str(value or "").strip()

        if "プライム" in text or "prime" in text.lower():
            return "Prime"
        if (
            "スタンダード" in text
            or "standard" in text.lower()
        ):
            return "Standard"
        if "グロース" in text or "growth" in text.lower():
            return "Growth"

        return None

    @staticmethod
    def _is_domestic_common_stock(
        category: str,
    ) -> bool:
        normalized = category.lower()

        excluded = (
            "etf",
            "reit",
            "インフラ",
            "出資証券",
            "外国株",
            "foreign",
            "優先",
            "preferred",
        )
        if any(
            marker in normalized
            for marker in excluded
        ):
            return False

        domestic_markers = (
            "内国株",
            "domestic",
        )
        return any(
            marker in normalized
            for marker in domestic_markers
        )

    def _current_time(self) -> datetime:
        value = self.now_provider()
        if value.tzinfo is None:
            raise ValueError(
                "現在日時にはタイムゾーンが必要です。"
            )
        return value.astimezone(timezone.utc)
